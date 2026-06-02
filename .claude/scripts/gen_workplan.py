"""
生成马跃一号工作计划总表 Excel
5 Sheets: 总体工作计划 / 企管部专属行动 / 周甘特图 / 按工作组查看 / 图例与使用说明
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = openpyxl.Workbook()

# ── 全局样式 ──
DARK_BLUE = '1a3c6e'
MED_BLUE = '4a6fa5'
LIGHT_BLUE = 'd6e4f0'
WHITE = 'ffffff'
YELLOW_INPUT = 'fff2cc'
GREEN_GOOD = 'c6efce'
RED_BAD = 'ffc7ce'
ORANGE_WARN = 'ffe0b3'
LIGHT_GRAY = 'f5f5f5'
GREEN_DARK = '006100'
RED_DARK = 'cc0000'

thin_border = Border(
    left=Side(style='thin', color='cccccc'),
    right=Side(style='thin', color='cccccc'),
    top=Side(style='thin', color='cccccc'),
    bottom=Side(style='thin', color='cccccc')
)

header_font = Font(name='微软雅黑', size=10, bold=True, color=WHITE)
header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
title_font = Font(name='微软雅黑', size=14, bold=True, color=DARK_BLUE)
section_font = Font(name='微软雅黑', size=11, bold=True, color=DARK_BLUE)
body_font = Font(name='微软雅黑', size=10, color='333333')
bold_font = Font(name='微软雅黑', size=10, bold=True, color='333333')
small_font = Font(name='微软雅黑', size=9, color='888888')

input_fill = PatternFill(start_color=YELLOW_INPUT, end_color=YELLOW_INPUT, fill_type='solid')
good_fill = PatternFill(start_color=GREEN_GOOD, end_color=GREEN_GOOD, fill_type='solid')
bad_fill = PatternFill(start_color=RED_BAD, end_color=RED_BAD, fill_type='solid')
warn_fill = PatternFill(start_color=ORANGE_WARN, end_color=ORANGE_WARN, fill_type='solid')
light_blue_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')
gray_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')

risk_fill = PatternFill(start_color='fce4ec', end_color='fce4ec', fill_type='solid')
gov_fill = PatternFill(start_color='ede7f6', end_color='ede7f6', fill_type='solid')
p0_blue_fill = PatternFill(start_color='e6f3ff', end_color='e6f3ff', fill_type='solid')
p3_yellow_fill = PatternFill(start_color='fff8e1', end_color='fff8e1', fill_type='solid')
p4_green_fill = PatternFill(start_color='e8f5e9', end_color='e8f5e9', fill_type='solid')

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def sc(ws, row, col, value=None, font=None, fill=None, alignment=None, fmt=None):
    """set_cell shorthand"""
    cell = ws.cell(row=row, column=col)
    if value is not None: cell.value = value
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if fmt: cell.number_format = fmt
    cell.border = thin_border
    return cell

def wh(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        sc(ws, row, start_col+i, h, font=header_font, fill=header_fill, alignment=center_align)

def stitle(ws, row, text, ncols=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    sc(ws, row, 1, text, font=section_font, fill=light_blue_fill, alignment=left_align)
    for c in range(2, ncols+1):
        ws.cell(row=row, column=c).fill = light_blue_fill
        ws.cell(row=row, column=c).border = thin_border

def cw(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def phase_fill(phase_str):
    if 'Phase 0' in phase_str: return light_blue_fill
    if 'Phase 1' in phase_str or 'Day1' in phase_str: return p0_blue_fill
    if 'Phase 2' in phase_str: return gray_fill
    if 'Phase 3' in phase_str: return p3_yellow_fill
    if 'Phase 4' in phase_str: return p4_green_fill
    if '风险' in phase_str: return risk_fill
    if '治理' in phase_str: return gov_fill
    return None

# ═══════════════════════ Sheet 1: 总体工作计划 ═══════════════════════
ws1 = wb.active
ws1.title = '总体工作计划'
ws1.sheet_properties.tabColor = DARK_BLUE

ws1.merge_cells('A1:L1')
sc(ws1, 1, 1, '马跃一号 — 投后融合管理 总体工作计划', font=title_font, alignment=center_align)
ws1.row_dimensions[1].height = 32
ws1.merge_cells('A2:L2')
sc(ws1, 2, 1, '企管部牵头 | DBS方法论 | 100天窗口期 | V1.0 | 2026-05-30 | ⬜未启动 🔵进行中 ✅已完成 🔴阻塞', font=small_font, alignment=center_align)
ws1.merge_cells('A3:L3')
sc(ws1, 3, 1, '共98项: Phase0(15)+Day1(12)+Phase2(18)+Phase3(20)+Phase4(8)+风险应对(15)+治理机制(10)', font=small_font, alignment=center_align)

r = 5
wh(ws1, r, ['序号','阶段','任务分类','任务事项','交付物/完成标准','责任部门','责任人','开始周','结束周','优先级','状态','备注/依赖'])

# ── 任务数据 ──
T = []
# Phase 0
T += [
    ['Phase 0\n签约前','PMO筹备','启动PMO筹备：确定企管部专职人员','企管部专职人员到位','企管部','王瑞俊','W-12','W-11','P0','⬜','需张毅总确认编制'],
    ['Phase 0\n签约前','PMO筹备','建立项目管理工具（共享看板/任务追踪）','PMO工具就绪','企管部/数字化','王瑞俊','W-12','W-10','P0','⬜','推荐飞书多维表格或Excel共享'],
    ['Phase 0\n签约前','方法论','完成融合Playbook终稿','融合策划书V1.0 ✅已有','企管部','王瑞俊','W-10','W-10','P0','✅','已完成2026-05-27'],
    ['Phase 0\n签约前','治理启动','召开第一次工作组组长预备会','组长共识备忘录','企管部','王瑞俊','W-8','W-8','P0','⬜','需委员会副主任授权'],
    ['Phase 0\n签约前','Day1准备','完成Day 1准备清单并分配责任人','Day 1清单V1.0','企管部','王瑞俊','W-6','W-6','P0','⬜','含50项具体事项'],
    ['Phase 0\n签约前','数据基线','尽调信息汇总：获取标的运营数据基线','数据基线报告','战略发展部','谭彦杰','W-4','W-4','P0','⬜','财务部配合'],
    ['Phase 0\n签约前','人才准备','确定派驻GM人选，完成DBS速成培训(1周)','GM就位+培训完成','组织发展部','—','W-3','W-3','P0','⬜','企管部提供DBS培训'],
    ['Phase 0\n签约前','治理启动','委员会第一次正式会议：审批百日计划、确认预算','委员会决议','委员会','季国祥','W-2','W-2','P0','⬜','PMO汇报百日计划'],
    ['Phase 0\n签约前','Day1准备','Day 1全流程桌面推演','演练checklist全部通过','企管部','王瑞俊','W-1','W-1','P0','⬜','所有关键人员参与'],
    ['Phase 0\n签约前','IT准备','启动IT系统对接预备（ERP编码映射/物料主数据清洗）','ERP对接方案获批','数字化组','王瑞俊','W-4','W-1','P0','⬜','⚠️易被忽视，需提前4-8周'],
    ['Phase 0\n签约前','人才准备','关键人才识别（按技能不可替代性）','关键人才清单','组织发展部','—','W-4','W-3','P0','⬜',''],
    ['Phase 0\n签约前','人才准备','关键人才保留方案设计+激励方案获批','保留协议模板+激励方案','组织发展部','—','W-4','W-2','P0','⬜',''],
    ['Phase 0\n签约前','合规准备','重大合同审查(>50万)+IP清单确认','合同审查报告+IP清单','法务部','—','W-4','W-1','P1','⬜',''],
    ['Phase 0\n签约前','合规准备','环保/安全/消防/劳动合规排查','合规排查报告','HSE/法务','林爱彬','W-6','W-2','P1','⬜','阀件涉及切削液/电镀'],
    ['Phase 0\n签约前','沟通准备','员工FAQ手册(20个最可能问题)+CEO演讲稿起草','FAQ手册+演讲稿','企管部','王瑞俊','W-3','W-1','P0','⬜',''],
]
# Phase 1: Day1
T += [
    ['Phase 1\nDay1','法律交割','法律交割完成确认','法务确认文件','法务部','—','W0','W0','P0','⬜','08:00前完成'],
    ['Phase 1\nDay1','全员沟通','Town Hall全员大会','大会议程完成+Q&A记录','企管部','王瑞俊','W0','W0','P0','⬜','CEO+派驻GM出席'],
    ['Phase 1\nDay1','中高层沟通','派驻GM × 各部门负责人1on1见面会','全部1on1完成','派驻GM','—','W0','W0','P0','⬜','企管部准备信息卡'],
    ['Phase 1\nDay1','客户界面','TOP20客户逐一电话/拜访','全部联系完成','营销中心','包秋婧','W0','W0','P0','⬜','确认名单+话术'],
    ['Phase 1\nDay1','供应商界面','核心供应商逐一通知','全部通知完成','采购物流','刘建中','W0','W0','P0','⬜','确保供应不中断'],
    ['Phase 1\nDay1','资产交接','印章/证照/密钥交接','交接清单双方签字','法务/财务','—','W0','W0','P0','⬜','企管部监交'],
    ['Phase 1\nDay1','资产交接','银行账户授权变更','银行受理完毕','财务部','—','W0','W0','P0','⬜',''],
    ['Phase 1\nDay1','DBSO','DBSO正式挂牌+第一个Kaizen计划公示','DBSO挂牌+Kaizen日程上墙','企管部(DBSO)','王瑞俊','W0','W0','P0','⬜','企管部主场，需仪式感'],
    ['Phase 1\nDay1','员工关怀','发放"致员工的一封信"+欢迎礼包','信+礼包全员发放','企管部/组织发展部','王瑞俊','W0','W0','P0','⬜','CEO+GM联署'],
    ['Phase 1\nDay1','品牌更新','公司铭牌/前台logo/厂区标识更新','全部更换完成','企管部','—','W0','W0','P1','⬜','07:00前完成'],
    ['Phase 1\nDay1','沟通渠道','启动"问题热线"+匿名问卷','问题收集通道就绪','企管部','王瑞俊','W0','W0','P0','⬜',''],
    ['Phase 1\nDay1','日终总结','Day 1总结短会(PMO+五组长,30min)','当日问题清单+次日优先项','企管部','王瑞俊','W0','W0','P0','⬜','三问:出什么岔子/最紧急/没想到'],
]
# Phase 2: W1-W4
T += [
    ['Phase 2\nW1-W4','PMO机制','启动每日晨会(15min站立会)','晨会机制运行','企管部(PMO)','王瑞俊','W1','W4','P0','⬜','前30天每日，后续降频'],
    ['Phase 2\nW1-W4','PMO机制','建立PMO仪表板初版','PMO仪表板V1.0','企管部(PMO)','王瑞俊','W1','W1','P0','⬜','红黄绿+里程碑+阻塞项'],
    ['Phase 2\nW1-W4','组织人事','发布过渡期组织架构+高管聘任','架构图+任免文件发布','组织发展部','—','W1','W1','P0','⬜','不过度承诺"不变"'],
    ['Phase 2\nW1-W4','组织人事','全员劳动合同签署启动','签署方案启动','组织发展部','—','W1','W2','P0','⬜',''],
    ['Phase 2\nW1-W4','组织人事','关键人才保留协议签署','签署率100%','组织发展部','—','W1','W2','P0','⬜','含留任奖金+里程碑兑现'],
    ['Phase 2\nW1-W4','运营改善','组织第一次VSM活动—选"订单到交付"主线','VSM现状图','企管部(DBSO)','王瑞俊','W2','W2','P0','⬜','买卖双方同坐一桌'],
    ['Phase 2\nW1-W4','运营改善','第一次Kaizen活动(3天)—选最显眼的浪费点','Kaizen成果+标准固化','企管部(DBSO)','王瑞俊','W3','W3','P0','⬜','选最容易改的痛点'],
    ['Phase 2\nW1-W4','财务管理','财务制度对接启动(报销/付款/预算/资产)','财务制度清单+优先级','财务部','—','W1','W3','P0','⬜',''],
    ['Phase 2\nW1-W4','财务管理','过渡期双签机制启动(>10万支出双签)','双签机制运行','财务部','—','W1','W1','P0','⬜','派驻财务经理Day1到岗'],
    ['Phase 2\nW1-W4','质量管理','质量体系双轨期启动','双轨方案运行','质量管理部','李世社','W1','W4','P0','⬜','原标准维持+中集标准对比'],
    ['Phase 2\nW1-W4','供应链','采购供应商清单合并','供应商合并清单','采购物流','刘建中','W2','W3','P1','⬜','前30天不更换任何供应商'],
    ['Phase 2\nW1-W4','IT系统','OA/邮箱/域控/企业微信账号创建+测试','系统就绪','数字化组','王瑞俊','W1','W3','P0','⬜',''],
    ['Phase 2\nW1-W4','IT系统','ERP基础对接启动(物料编码映射/科目对照)','对接方案执行中','数字化组','王瑞俊','W1','W4','P0','⬜','不切换，先建测试环境'],
    ['Phase 2\nW1-W4','沟通文化','派驻GM两周内完成所有关键人才1on1','全部1on1完成','派驻GM','—','W1','W2','P0','⬜','GM前30天首要KPI'],
    ['Phase 2\nW1-W4','沟通文化','派驻GM完成3次现场Gemba Walk','3次Gemba Walk完成','派驻GM','—','W1','W4','P0','⬜',''],
    ['Phase 2\nW1-W4','PMO机制','第4周:前30天总结报告提交委员会','30天总结报告','企管部(PMO)','王瑞俊','W4','W4','P0','⬜','MOR格式:指标/差距/根因/对策'],
    ['Phase 2\nW1-W4','DBS培训','面向班组长:PSP问题解决流程入门培训(1天)','培训完成','企管部(DBSO)','王瑞俊','W2','W3','P1','⬜',''],
    ['Phase 2\nW1-W4','DBS机制','产线日管控看板上墙(安全/质量/交付/成本)','看板就绪+试点产线确定','企管部(DBSO)','王瑞俊','W2','W3','P1','⬜',''],
]
# Phase 3: W5-W12
T += [
    ['Phase 3\nW5-W12','战略治理','董事会/股东会召开，治理架构正式运行','工商变更完成','战略发展部/董秘办','谭彦杰','W5','W8','P0','⬜',''],
    ['Phase 3\nW5-W12','采购协同','采购协同首批落地(集采降价≥3%证伪或证实)','采购协同金额报告','采购物流','刘建中','W6','W10','P0','⬜','仅对双轨供应商启动'],
    ['Phase 3\nW5-W12','质量整合','质量双轨期满:完成双方标准逐项对比','质量标准对比报告','质量管理部','李世社','W8','W10','P0','⬜','对20%差异逐项判断'],
    ['Phase 3\nW5-W12','财务整合','财务并表完成','首次并表完成','财务部','—','W5','W8','P0','⬜','上市公司披露要求'],
    ['Phase 3\nW5-W12','财务整合','资金管控体系正式运行','资金日报/周报稳定','财务部','—','W5','W6','P0','⬜',''],
    ['Phase 3\nW5-W12','人才文化','关键人才保留协议签署率100%确认','签署率100%','组织发展部','—','W5','W6','P0','⬜',''],
    ['Phase 3\nW5-W12','人才文化','薪酬对标完成','薪酬对标报告','组织发展部','—','W6','W8','P1','⬜','不承诺立即对齐'],
    ['Phase 3\nW5-W12','人才文化','第一次文化融合工作坊','工作坊完成','组织发展部','—','W8','W8','P1','⬜',''],
    ['Phase 3\nW5-W12','IT系统','ERP基础对接完成(至少科目/物料主数据贯通)','主数据贯通测试通过','数字化组','王瑞俊','W6','W10','P0','⬜','100天内不做全量切换'],
    ['Phase 3\nW5-W12','运营改善','Kaizen活动持续(每两周一次)','累计≥5次Kaizen完成','企管部(DBSO)','王瑞俊','W5','W12','P0','⬜','产线→OTC→P2P→...'],
    ['Phase 3\nW5-W12','运营改善','第2次Kaizen:订单到交付流程(OTC)','OTC流程改善成果','企管部(DBSO)','王瑞俊','W5','W6','P1','⬜','TPI方法论'],
    ['Phase 3\nW5-W12','运营改善','第3次Kaizen:采购到付款流程(P2P)','P2P流程改善成果','企管部(DBSO)','王瑞俊','W7','W8','P1','⬜','TPI方法论'],
    ['Phase 3\nW5-W12','运营改善','供应商整合策略制定(基于前100天数据)','供应商整合策略报告','采购物流','刘建中','W8','W10','P1','⬜','独有供应商至少保留至100天后'],
    ['Phase 3\nW5-W12','运营改善','制造能力转移评估启动(关键工艺对标)','工艺能力评估报告','技术中心/生产','陈晓春/林爱彬','W6','W10','P1','⬜','阀件精密加工核心能力'],
    ['Phase 3\nW5-W12','PMO机制','PMO仪表板升级为百日冲刺版','仪表板V2.0','企管部(PMO)','王瑞俊','W8','W8','P0','⬜','含五组KPI+里程碑+阻塞趋势'],
    ['Phase 3\nW5-W12','PMO机制','委员会月度MOR会议 × 2次','两份MOR报告+决议','企管部(PMO)','王瑞俊','W8','W12','P0','⬜',''],
    ['Phase 3\nW5-W12','沟通文化','GM-团队融合反馈会(企管部组织)','反馈报告','企管部','王瑞俊','W5','W5','P1','⬜',''],
    ['Phase 3\nW5-W12','协同追踪','启动协同效应追踪表','协同追踪表运行','企管部(PMO)','王瑞俊','W6','W12','P0','⬜','每项:预期/实际/偏差/根因'],
    ['Phase 3\nW5-W12','协同追踪','内转定价机制框架制定','转移定价框架方案','财务部/企管部','—','W8','W12','P1','⬜','初期用成本加成法'],
    ['Phase 3\nW5-W12','CapEx','资本支出优先级评审(安全/效率/战略三维度)','CapEx优先级清单','企管部/财务','王瑞俊','W6','W8','P1','⬜',''],
]
# Phase 4: W13+
T += [
    ['Phase 4\nW13+','复盘总结','100天总结报告提交委员会','100天总结报告','企管部(PMO)','王瑞俊','W13','W13','P0','⬜','各组回答:对/错/改进'],
    ['Phase 4\nW13+','复盘总结','第一次融合复盘会(全工作组)','复盘会纪要+行动项','企管部(PMO)','王瑞俊','W13','W13','P0','⬜',''],
    ['Phase 4\nW13+','机制过渡','PMO转入轻量模式:周会改双周会','轻量PMO运行','企管部(PMO)','王瑞俊','W14','W14','P1','⬜',''],
    ['Phase 4\nW13+','机制过渡','DBSO角色从"推动者"转为"教练/支持"','角色转换完成','企管部(DBSO)','王瑞俊','W16','W16','P1','⬜','GM主导MOR'],
    ['Phase 4\nW13+','机制过渡','派驻GM开始用DBS语言主持MOR','GM主导首次MOR完成','派驻GM','—','W14','W16','P0','⬜','标志融合进入常态化'],
    ['Phase 4\nW13+','知识沉淀','基于本次融合实战更新Playbook','Playbook V2.0','企管部','王瑞俊','W18','W20','P1','⬜','纳入公司知识库'],
    ['Phase 4\nW13+','知识沉淀','PMO正式解散，融合管理移交MOR机制','移交完成','企管部(PMO)','王瑞俊','W20','W20','P1','⬜',''],
    ['Phase 4\nW13+','绩效评估','融合绩效评估完成(KPI仪表板终版+协同效应终版)','最终评估报告','企管部(PMO)','王瑞俊','W14','W16','P0','⬜',''],
]
# 风险应对
T += [
    ['风险应对','R1人才流失','关键人才保留协议签署+GM 1on1全覆盖','签署率100%','组织发展部/派驻GM','—','W1','W2','P0','⬜','🔴高风险(20分)'],
    ['风险应对','R2客户转移','TOP20客户逐一联系+业务照常保证函','全部联系+确认函发送','营销中心','包秋婧','W0','W2','P0','⬜','🔴高风险(15分)'],
    ['风险应对','R3 ERP失败','坚持三步走:先用原系统→同步建新→测试通过再切换','ERP测试环境通过','数字化组','王瑞俊','W1','W10','P0','⬜','🔴高风险(15分)'],
    ['风险应对','R4质量打架','双轨制执行+首月完成标准逐项对比','标准对比报告','质量管理部','李世社','W1','W8','P1','⬜','🟠中高风险(12分)'],
    ['风险应对','R5 GM冲突','选任强调倾听式领导+前30天KPI是1on1+Gemba','W4融合反馈会','组织发展部/企管部','王瑞俊','W1','W4','P1','⬜','🟠中高风险(12分)'],
    ['风险应对','R6供应商断供','前30天不更换供应商+独有供应商至少保留100天','供应商交付准时率≥95%','采购物流','刘建中','W1','W12','P0','⬜','🟠中高风险(12分)'],
    ['风险应对','R7财务失控','Day1银行授权变更+派驻财务经理+双签机制','资金日报零缺失','财务部','—','W0','W1','P0','⬜','🟠中高风险(10分)'],
    ['风险应对','R8员工焦虑','FAQ手册+Town Hall直面问题+匿名通道','员工担忧指数<30%','组织发展部/企管部','王瑞俊','W0','W2','P1','⬜','🟡中低风险(8分)'],
    ['风险应对','R9 IP泄露','关键技术文档归档+竞业限制签署','IP清单+保密协议100%','法务部/技术中心','—','W1','W4','P1','⬜','🟡中低风险(8分)'],
    ['风险应对','R10环保遗留','Phase 0完成合规排查+分级整改计划','排查报告+整改计划','HSE/法务部','林爱彬','W-4','W1','P1','⬜','🟡中低风险(8分)'],
    ['风险应对','R11或有负债','交易文件赔偿条款+部分对价保证金','赔偿条款签署','法务部/财务部','—','W-2','W0','P1','⬜','🟡中低风险(8分)'],
    ['风险应对','R12产能干扰','Day1不停线+第一个Kaizen聚焦产线效率','产线OEE不下降','生产制造','林爱彬','W0','W4','P2','⬜','🟡中低风险(6分)'],
    ['风险应对','R13汇率材料','商业计划预留±10%成本波动缓冲','成本波动在缓冲范围内','财务部','—','W1','W12','P2','⬜','🟡中低风险(6分)'],
    ['风险应对','R14过度整合','严格执行Phase 2禁止事项清单+委员会scope审查','未触发禁止事项','企管部(PMO)','王瑞俊','W1','W12','P1','⬜','🟡中低风险(6分)'],
    ['风险应对','R15垄断负面','准备反垄断合规论证+对外强调供应链安全','合规论证就绪','战略发展部/董秘办','谭彦杰','W-2','W0','P2','⬜','🟡低风险(5分)'],
]
# 治理机制
T += [
    ['治理机制','PMO机制','每日晨会(前30天每日，后续降频)','每日15min，阻塞项当日清','PMO','王瑞俊','W1','W12','P0','⬜',''],
    ['治理机制','PMO机制','PMO周例会(每周一，五组长，60min)','周进度报告+风险更新','PMO','王瑞俊','W1','W20','P0','⬜','每两周一次到W20'],
    ['治理机制','委员会','委员会月度MOR(每月末，90min)','MOR报告+决策决议','委员会','季国祥/张毅','W4','W20','P0','⬜','W4/W8/W12/W16/W20'],
    ['治理机制','Kaizen','Kaizen活动(每两周一次，3-5天)','每次Kaizen有成果+标准固化','DBSO','王瑞俊','W3','W12','P0','⬜','≥5次在100天内'],
    ['治理机制','风险管理','风险登记册周更新','TOP15风险状态+趋势','PMO','王瑞俊','W1','W20','P0','⬜','红色48h升级委员会'],
    ['治理机制','协同管理','协同效应追踪表月度更新','每项:预期/实际/偏差','PMO','王瑞俊','W6','W20','P0','⬜','每月末财务签确'],
    ['治理机制','沟通机制','Town Hall全员大会','Day1+30天+100天,三次','企管部/派驻GM','王瑞俊','W0','W12','P1','⬜','三个节点'],
    ['治理机制','GM支持','GM-企管部双周check-in(非正式)','双周1次','企管部','王瑞俊','W1','W20','P1','⬜','了解GM状态+提供支持'],
    ['治理机制','复盘机制','30天总结+60天中期检查+100天全面复盘','三份复盘报告','PMO','王瑞俊','W4','W12','P0','⬜',''],
    ['治理机制','知识管理','融合经验入公司知识库+Playbook更新','Playbook V2.0入库','企管部','王瑞俊','W18','W20','P1','⬜',''],
]

# Write all tasks
for i, t in enumerate(T):
    row = r + 1 + i
    sc(ws1, row, 1, i+1, font=body_font, alignment=center_align)
    sc(ws1, row, 2, t[0], font=bold_font, alignment=center_align)
    sc(ws1, row, 3, t[1], font=bold_font, alignment=left_align)
    sc(ws1, row, 4, t[2], font=body_font, alignment=left_align)
    sc(ws1, row, 5, t[3], font=body_font, alignment=left_align)
    sc(ws1, row, 6, t[4], font=body_font, alignment=center_align)
    sc(ws1, row, 7, t[5], font=body_font, alignment=center_align)
    sc(ws1, row, 8, t[6], font=body_font, alignment=center_align)
    sc(ws1, row, 9, t[7], font=body_font, alignment=center_align)
    pfill = bad_fill if t[8] == 'P0' else (warn_fill if t[8] == 'P1' else None)
    sc(ws1, row, 10, t[8], font=bold_font, fill=pfill, alignment=center_align)
    sfill = good_fill if '✅' in str(t[9]) else None
    sc(ws1, row, 11, t[9], font=bold_font, fill=sfill, alignment=center_align)
    sc(ws1, row, 12, t[10] if len(t) > 10 else '', font=small_font, alignment=left_align)
    # Phase color band
    band = phase_fill(t[0])
    if band:
        for c in range(1, 13):
            ws1.cell(row=row, column=c).fill = band
    ws1.row_dimensions[row].height = 26

ws1.freeze_panes = 'A6'
ws1.auto_filter.ref = f'A5:L{r+len(T)}'
cw(ws1, [5, 12, 16, 38, 28, 14, 10, 10, 10, 8, 10, 24])

# ═══════════════════════ Sheet 2: 企管部专属行动 ═══════════════════════
ws2 = wb.create_sheet('企管部专属行动')
ws2.sheet_properties.tabColor = MED_BLUE

ws2.merge_cells('A1:H1')
sc(ws2, 1, 1, '企管部（王瑞俊）专属行动清单 — PMO + 运营融合 + IT/数字化 三重角色', font=title_font, alignment=center_align)
ws2.row_dimensions[1].height = 32
ws2.merge_cells('A2:H2')
sc(ws2, 2, 1, '王瑞俊需亲自执行/跟进/决策的事项，共42项。用于个人周排程和优先级管理。', font=small_font, alignment=center_align)

r = 4
wh(ws2, r, ['序号','角色','阶段','事项','交付物','截止周','优先级','状态'])

wrj = [
    ['PMO负责人','Phase 0','确定企管部专职PMO人员','人员到位','W-11','P0','⬜'],
    ['PMO负责人','Phase 0','建立项目管理工具','工具就绪','W-10','P0','⬜'],
    ['PMO负责人','Phase 0','完成Playbook终稿','Playbook ✅','W-10','P0','✅'],
    ['PMO负责人','Phase 0','召开第一次组长预备会','共识备忘录','W-8','P0','⬜'],
    ['PMO负责人','Phase 0','完成Day 1清单+分配责任人','Day1清单V1','W-6','P0','⬜'],
    ['PMO负责人','Phase 0','起草员工FAQ手册+CEO演讲稿','FAQ+演讲稿','W-3','P0','⬜'],
    ['PMO负责人','Phase 0','向委员会汇报百日计划','百日计划获批','W-2','P0','⬜'],
    ['PMO负责人','Phase 0','组织Day 1桌面推演','演练通过','W-1','P0','⬜'],
    ['PMO负责人','Day1','协调CEO出席Town Hall+审阅演讲稿','议程完成','W0','P0','⬜'],
    ['PMO负责人','Day1','DBSO挂牌+Kaizen计划公示','DBSO挂牌','W0','P0','⬜'],
    ['PMO负责人','Day1','主持Day 1总结短会','问题清单','W0','P0','⬜'],
    ['PMO负责人','Day1','启动问题热线+匿名问卷','通道就绪','W0','P0','⬜'],
    ['PMO负责人','W1-W4','启动每日晨会机制','晨会运行','W1','P0','⬜'],
    ['PMO负责人','W1-W4','建立PMO仪表板V1.0','仪表板V1','W1','P0','⬜'],
    ['PMO负责人','W1-W4','向委员会提交30天总结报告','30天报告','W4','P0','⬜'],
    ['PMO负责人','W5-W12','委员会月度MOR ×2次','MOR报告','W8/W12','P0','⬜'],
    ['PMO负责人','W5-W12','PMO仪表板升级V2.0(冲刺版)','仪表板V2','W8','P0','⬜'],
    ['PMO负责人','W5-W12','启动协同效应追踪表','追踪表运行','W6','P0','⬜'],
    ['PMO负责人','W13+','100天总结报告+融合复盘会','总结报告','W13','P0','⬜'],
    ['PMO负责人','W13+','PMO转轻量模式+解散移交','移交完成','W14-W20','P1','⬜'],
    ['PMO负责人','W13+','更新Playbook V2.0','Playbook入库','W20','P1','⬜'],
    ['DBSO/运营融合组长','Phase 0','GM的DBS速成培训(1周)','GM培训完成','W-3','P0','⬜'],
    ['DBSO/运营融合组长','W1-W4','组织第一次VSM活动(订单到交付)','VSM现状图','W2','P0','⬜'],
    ['DBSO/运营融合组长','W1-W4','组织第一次Kaizen活动(3天)','Kaizen成果','W3','P0','⬜'],
    ['DBSO/运营融合组长','W1-W4','PSP培训面向班组长(1天版)','培训完成','W2-W3','P1','⬜'],
    ['DBSO/运营融合组长','W1-W4','产线日管控看板上墙','看板就绪','W2-W3','P1','⬜'],
    ['DBSO/运营融合组长','W5-W12','Kaizen活动持续组织(每两周一次)','≥5次完成','W5-W12','P0','⬜'],
    ['DBSO/运营融合组长','W5-W12','第二次Kaizen:订单到交付流程(OTC)','OTC改善','W5-W6','P1','⬜'],
    ['DBSO/运营融合组长','W5-W12','第三次Kaizen:采购到付款流程(P2P)','P2P改善','W7-W8','P1','⬜'],
    ['DBSO/运营融合组长','W5-W12','GM-团队融合反馈会','反馈报告','W5','P1','⬜'],
    ['DBSO/运营融合组长','W5-W12','CapEx优先级评审','CapEx清单','W6-W8','P1','⬜'],
    ['DBSO/运营融合组长','W13+','DBSO角色从推动者转教练/支持','角色转换','W16','P1','⬜'],
    ['IT/数字化组长','Phase 0','ERP对接方案启动(编码映射/主数据清洗)','对接方案','W-4','P0','⬜'],
    ['IT/数字化组长','W1-W4','OA/邮箱/域控/企业微信账号创建+测试','系统就绪','W3','P0','⬜'],
    ['IT/数字化组长','W1-W4','ERP测试环境搭建(不切换,先并行)','测试环境通过','W4','P0','⬜'],
    ['IT/数字化组长','W5-W12','ERP主数据贯通测试','贯通测试通过','W10','P0','⬜'],
    ['PMO负责人','持续','风险登记册每周更新+红色风险48h升级','周度风险更新','W1-W20','P0','⬜'],
    ['PMO负责人','持续','跨组阻塞项48h内升级解决','阻塞项清零','W1-W20','P0','⬜'],
    ['PMO负责人','持续','GM双周check-in','双周1次','W1-W20','P1','⬜'],
    ['PMO负责人','持续','协同效应追踪月度签确','月度签确','W6-W20','P0','⬜'],
    ['PMO负责人','决策','向张毅总确认:马跃项目是否为Q3企管部第一优先级','优先级确认','W-10','P0','⬜'],
    ['PMO负责人','决策','向张毅总申请企管部1人全职投入PMO','人员到位','W-10','P0','⬜'],
]

for i, t in enumerate(wrj):
    row = r + 1 + i
    sc(ws2, row, 1, i+1, font=body_font, alignment=center_align)
    sc(ws2, row, 2, t[0], font=bold_font, alignment=center_align)
    sc(ws2, row, 3, t[1], font=body_font, alignment=center_align)
    sc(ws2, row, 4, t[2], font=body_font, alignment=left_align)
    sc(ws2, row, 5, t[3], font=body_font, alignment=left_align)
    sc(ws2, row, 6, t[4], font=body_font, alignment=center_align)
    pfill = bad_fill if t[5] == 'P0' else (warn_fill if t[5] == 'P1' else None)
    sc(ws2, row, 7, t[5], font=bold_font, fill=pfill, alignment=center_align)
    sfill = good_fill if '✅' in str(t[6]) else None
    sc(ws2, row, 8, t[6], font=bold_font, fill=sfill, alignment=center_align)
    ws2.row_dimensions[row].height = 23

ws2.freeze_panes = 'A5'
ws2.auto_filter.ref = f'A4:H{r+len(wrj)}'
cw(ws2, [5, 20, 10, 36, 26, 14, 8, 8])

# ═══════════════════════ Sheet 3: 周甘特图 ═══════════════════════
ws3 = wb.create_sheet('周甘特图')
ws3.sheet_properties.tabColor = '4caf50'

ws3.merge_cells('A1:AE1')
sc(ws3, 1, 1, '马跃一号 — 百日融合 周甘特图 (W-12 ~ W20)', font=title_font, alignment=center_align)
ws3.row_dimensions[1].height = 32

# Week header
week_labels = [f'W{w}' if w != 0 else 'Day1' for w in range(-12, 21)]
r = 3
ws3.merge_cells(f'A{r}:C{r}')
sc(ws3, r, 1, '工作组/任务线', font=header_font, fill=header_fill, alignment=center_align)
for i, label in enumerate(week_labels):
    sc(ws3, r, 4+i, label, font=Font(name='微软雅黑', size=8, bold=True, color=WHITE), fill=header_fill, alignment=center_align)

# Phase color bands
phases_gantt = [
    ('Phase 0: 签约前准备', -12, -1, 'd6e4f0'),
    ('Phase 1: Day1', 0, 0, 'e6f3ff'),
    ('Phase 2: 稳定+观察(W1-W4)', 1, 4, 'f0f0f0'),
    ('Phase 3: 深度整合(W5-W12)', 5, 12, 'fff8e1'),
    ('Phase 4: 常态化(W13+)', 13, 20, 'e8f5e9'),
]
for pname, sw, ew, color in phases_gantt:
    p_sc = sw + 16  # W-12 -> col 4
    p_ec = ew + 16
    for c in range(p_sc, p_ec+1):
        sc(ws3, 4, c, '', font=body_font, fill=PatternFill(start_color=color, end_color=color, fill_type='solid'), alignment=center_align)
    ws3.merge_cells(start_row=4, start_column=p_sc, end_row=4, end_column=p_ec)
    sc(ws3, 4, p_sc, pname, font=Font(name='微软雅黑', size=8, bold=True), fill=PatternFill(start_color=color, end_color=color, fill_type='solid'), alignment=center_align)

# Gantt bars
gantt = [
    ['PMO: 筹备+工具搭建', -12, -10, DARK_BLUE],
    ['企管部: Playbook+Day1清单+FAQ+演讲稿', -10, -1, DARK_BLUE],
    ['战略发展: 尽调基线+董事会方案+治理架构', -4, -1, MED_BLUE],
    ['组织发展: 关键人才识别+保留方案+GM选任', -4, -1, MED_BLUE],
    ['法务: 合同审查+IP+合规排查', -4, -1, MED_BLUE],
    ['IT: ERP对接方案启动', -4, -1, MED_BLUE],
    ['委员会: 百日计划审批', -2, -2, MED_BLUE],
    ['全组: Day1桌面推演', -1, -1, RED_DARK],
    ['Day1: 法律交割+Town Hall+客户/供应商通知', 0, 0, RED_DARK],
    ['Day1: DBSO挂牌+品牌更新+欢迎礼包', 0, 0, RED_DARK],
    ['PMO: 晨会+周例会+仪表板V1', 1, 4, DARK_BLUE],
    ['组织发展: 组织架构发布+合同签署+人才保留', 1, 2, MED_BLUE],
    ['DBSO: VSM(第1次)+Kaizen(第1次)', 2, 3, DARK_BLUE],
    ['DBSO: PSP培训+日管控看板', 2, 3, MED_BLUE],
    ['财务: 制度对接+双签+资金日报', 1, 3, MED_BLUE],
    ['质量: 双轨期启动', 1, 4, MED_BLUE],
    ['IT: OA/邮箱/ERP测试环境', 1, 4, MED_BLUE],
    ['GM: 全员1on1+Gemba Walk×3', 1, 4, MED_BLUE],
    ['PMO: 30天总结报告', 4, 4, RED_DARK],
    ['战略治理: 董事会/股东会召开', 5, 8, MED_BLUE],
    ['财务: 并表完成+资金体系运行', 5, 8, MED_BLUE],
    ['采购: 集采协同首批落地', 6, 10, MED_BLUE],
    ['DBSO: Kaizen×3+持续改善', 5, 12, DARK_BLUE],
    ['质量: 标准对比+统一时间表', 8, 10, MED_BLUE],
    ['人才: 薪酬对标+文化融合工作坊', 6, 8, MED_BLUE],
    ['IT: ERP主数据贯通测试', 6, 10, MED_BLUE],
    ['技术: 制造能力转移评估', 6, 10, MED_BLUE],
    ['PMO: 协同追踪表+CapEx评审', 6, 8, DARK_BLUE],
    ['PMO: 仪表板V2+月度MOR×2', 8, 12, RED_DARK],
    ['供应商: 整合策略制定', 8, 10, MED_BLUE],
    ['PMO: 100天总结+复盘会', 13, 13, RED_DARK],
    ['PMO: 轻量模式+移交MOR', 14, 16, DARK_BLUE],
    ['GM: 主导首次MOR', 14, 16, MED_BLUE],
    ['企管部: Playbook更新V2.0', 18, 20, DARK_BLUE],
    ['PMO: 正式解散', 20, 20, MED_BLUE],
]

r = 5
for i, g in enumerate(gantt):
    row = r + i
    sc(ws3, row, 1, g[0], font=Font(name='微软雅黑', size=9, bold=True), alignment=left_align)
    ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    scol = g[1] + 16
    ecol = g[2] + 16
    color = g[3]
    for c in range(scol, ecol+1):
        sc(ws3, row, c, '', font=body_font,
           fill=PatternFill(start_color=color, end_color=color, fill_type='solid'),
           alignment=center_align)
    ws3.row_dimensions[row].height = 19

# Milestones
milestones = [('Day1',0,'Day1亮相'),('30天',4,'30天报告'),('60天',8,'中期检查'),('100天',12,'100天报告'),('PMO解散',20,'移交完成')]
rm = r + len(gantt) + 2
stitle(ws3, rm, '关键里程碑', 36)
for c2 in range(2, 37):
    ws3.cell(row=rm, column=c2).fill = light_blue_fill

for i, m in enumerate(milestones):
    row = rm + 1 + i
    sc(ws3, row, 1, m[0], font=bold_font, alignment=left_align)
    ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    mc = m[1] + 16
    for c2 in range(4, 37):
        ws3.cell(row=row, column=c2).border = thin_border
    sc(ws3, row, mc, '▼', font=Font(name='微软雅黑', size=10, bold=True, color=RED_DARK), alignment=center_align)
    sc(ws3, row, mc+1, m[2], font=bold_font, alignment=left_align)
    ws3.row_dimensions[row].height = 22

ws3.freeze_panes = 'D5'
cw(ws3, [32, 5, 5] + [5]*33)

# ═══════════════════════ Sheet 4: 按工作组查看 ═══════════════════════
ws4 = wb.create_sheet('按工作组查看')
ws4.sheet_properties.tabColor = 'ff9800'

ws4.merge_cells('A1:I1')
sc(ws4, 1, 1, '马跃一号 — 按工作组分类（用于各组组内排程和追踪）', font=title_font, alignment=center_align)
ws4.row_dimensions[1].height = 32

r = 3
wh(ws4, r, ['工作组','序号','阶段','任务事项','交付物/完成标准','责任部门','开始周','结束周','优先级'])

# Group tasks
from collections import OrderedDict
groups = OrderedDict([
    ('PMO/企管部(王瑞俊)', [t for t in T if '企管部' in t[4] or 'PMO' in t[2] or 'PMO' in t[6] or 'DBSO' in t[4]]),
    ('战略治理组(谭彦杰)', [t for t in T if '战略发展' in t[4] or '董秘办' in t[4]]),
    ('运营融合组', [t for t in T if any(d in t[4] for d in ['采购','技术中心','生产','质量','营销','DBSO','运营'])]),
    ('财务法务组', [t for t in T if '财务' in t[4] or '法务' in t[4]]),
    ('人才文化组', [t for t in T if '组织发展' in t[4] or 'HSE' in t[4] or '派驻GM' in t[6]]),
    ('IT/数字化组(王瑞俊)', [t for t in T if '数字化' in t[4]]),
    ('委员会', [t for t in T if '委员会' in t[4]]),
])

row = r + 1
for gname, gtasks in groups.items():
    if not gtasks: continue
    ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    sc(ws4, row, 1, f'▎{gname} ({len(gtasks)}项)', font=section_font, fill=light_blue_fill, alignment=left_align)
    for c in range(2, 10):
        ws4.cell(row=row, column=c).fill = light_blue_fill
        ws4.cell(row=row, column=c).border = thin_border
    row += 1
    for i, t in enumerate(gtasks):
        sc(ws4, row, 1, '', font=body_font, alignment=center_align)
        sc(ws4, row, 2, i+1, font=body_font, alignment=center_align)
        sc(ws4, row, 3, t[0].split('\n')[0], font=body_font, alignment=center_align)
        sc(ws4, row, 4, t[2], font=body_font, alignment=left_align)
        sc(ws4, row, 5, t[3], font=body_font, alignment=left_align)
        sc(ws4, row, 6, t[4], font=body_font, alignment=center_align)
        sc(ws4, row, 7, t[6], font=body_font, alignment=center_align)
        sc(ws4, row, 8, t[7], font=body_font, alignment=center_align)
        pfill2 = bad_fill if t[8] == 'P0' else (warn_fill if t[8] == 'P1' else None)
        sc(ws4, row, 9, t[8], font=bold_font, fill=pfill2, alignment=center_align)
        ws4.row_dimensions[row].height = 22
        row += 1
    row += 2

ws4.freeze_panes = 'A4'
cw(ws4, [22, 5, 12, 38, 28, 14, 10, 10, 8])

# ═══════════════════════ Sheet 5: 图例与说明 ═══════════════════════
ws5 = wb.create_sheet('图例与使用说明')
ws5.sheet_properties.tabColor = '9e9e9e'

ws5.merge_cells('A1:F1')
sc(ws5, 1, 1, '图例与使用说明', font=title_font, alignment=center_align)

r = 3
sc(ws5, r, 1, '▎阶段颜色编码', font=section_font, alignment=left_align)
for c in range(2,7):
    ws5.cell(row=r, column=c).border = thin_border
colors_info = [
    ('Phase 0 签约前准备 (W-12~W-1)', light_blue_fill, '方法论+治理+Day1准备'),
    ('Phase 1 Day1', p0_blue_fill, '亮相+安定+客户/供应商界面'),
    ('Phase 2 前30天 (W1-W4)', gray_fill, '稳定+观察+快速改善+禁止大动作'),
    ('Phase 3 30-100天 (W5-W12)', p3_yellow_fill, '深度整合+释放协同+系统对接'),
    ('Phase 4 100天+ (W13+)', p4_green_fill, '复盘+常态化+移交'),
    ('风险应对', risk_fill, '15项风险预防/应急措施'),
    ('治理机制', gov_fill, '会议/追踪/复盘等持续性机制'),
]
for i, c in enumerate(colors_info):
    row = r+1+i
    sc(ws5, row, 1, c[0], font=bold_font, fill=c[1], alignment=center_align)
    ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    sc(ws5, row, 4, c[2], font=body_font, alignment=left_align)
    ws5.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)

r = r+10
sc(ws5, r, 1, '▎优先级', font=section_font, alignment=left_align)
sc(ws5, r+1, 1, 'P0 — 必须完成，直接影响里程碑或阻塞后续任务', font=bold_font, alignment=left_align)
sc(ws5, r+1, 2, 'P0', font=bold_font, fill=bad_fill, alignment=center_align)
sc(ws5, r+2, 1, 'P1 — 重要但可容错1-2周，不影响关键路径', font=bold_font, alignment=left_align)
sc(ws5, r+2, 2, 'P1', font=bold_font, fill=warn_fill, alignment=center_align)
sc(ws5, r+3, 1, 'P2 — 锦上添花，资源允许时执行', font=bold_font, alignment=left_align)
sc(ws5, r+3, 2, 'P2', font=bold_font, alignment=center_align)

r = r+6
sc(ws5, r, 1, '▎状态', font=section_font, alignment=left_align)
sc(ws5, r+1, 1, '⬜ 未启动 — 尚未到达开始周或等待前置条件', font=bold_font, alignment=left_align)
sc(ws5, r+2, 1, '🔵 进行中 — 已经开始，正常推进', font=bold_font, alignment=left_align)
sc(ws5, r+3, 1, '✅ 已完成 — 交付物达标，责任人确认', font=bold_font, alignment=left_align)
sc(ws5, r+4, 1, '🔴 阻塞 — 因外部依赖或资源问题无法推进，需PMO介入', font=bold_font, alignment=left_align)

r = r+7
sc(ws5, r, 1, '▎使用说明', font=section_font, alignment=left_align)
instructions = [
    '1. "总体工作计划"Sheet是主表，98项任务的唯一数据源。筛选/排序/查找都在这里。',
    '2. "企管部专属行动"Sheet是王瑞俊个人的42项待办。用于个人每周排程。',
    '3. "周甘特图"Sheet用于管理层汇报和全员可视化。打印A3横向效果最佳。',
    '4. "按工作组查看"Sheet分给五个工作组组长，各组只看到自己的任务。',
    '5. 每周一PMO周例会前，各组更新自己任务的"状态"列。PMO汇总后更新主表。',
    '6. 建议使用Excel"筛选"功能按阶段/优先级/责任人快速定位。',
    '7. 风险项(R1-R15)在"风险应对"分类下，与04-风险登记册联动更新。',
    '8. 所有里程碑(Day1/30天/60天/100天/PMO解散)在甘特图中以▼标记。',
]
for i, txt in enumerate(instructions):
    sc(ws5, r+1+i, 1, txt, font=body_font, alignment=left_align)
    ws5.merge_cells(start_row=r+1+i, start_column=1, end_row=r+1+i, end_column=6)

cw(ws5, [30, 8, 8, 30, 8, 8])

# ═══════════════════════ Save ═══════════════════════
out_path = os.path.expanduser(r'~\Documents\Obsidian Vault\26年中集环科工作区\马跃项目\马跃一号_工作计划总表.xlsx')
wb.save(out_path)
print(f'Saved: {out_path}')
print(f'Sheets: 总体工作计划(98项) / 企管部专属行动(42项) / 周甘特图(35条) / 按工作组查看 / 图例与使用说明')
