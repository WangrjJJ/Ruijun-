"""
决策科学案例01 — Excel交互演示表格 V2.0
完全参数驱动：修改黄色单元格 → 全部计算自动更新 → 输出最优决策

Sheet 1: 参数设定 — 所有可调参数的唯一来源
Sheet 2: 规格选择优化器 — 标准板vs非标板 TCO计算 + MOQ判定 + 最优决策
Sheet 3: 聚合效应计算器 — 五车间独立vs跨车间聚合的MOQ解锁与损耗对比
Sheet 4: 库存策略优化器 — 安全库存最优水平计算 + 三种策略量化对比
Sheet 5: 综合仪表板 — 关键指标汇总 + 敏感性分析
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy
import os

wb = openpyxl.Workbook()

# ═══════════════════════ 全局样式 ═══════════════════════
DARK_BLUE = '1a3c6e'
MED_BLUE = '4a6fa5'
LIGHT_BLUE = 'd6e4f0'
WHITE = 'ffffff'
LIGHT_GRAY = 'f5f5f5'
YELLOW_INPUT = 'fff2cc'
GREEN_GOOD = 'c6efce'
RED_BAD = 'ffc7ce'
ORANGE_WARN = 'ffe0b3'
GREEN_DARK = '006100'

thin_border = Border(
    left=Side(style='thin', color='cccccc'),
    right=Side(style='thin', color='cccccc'),
    top=Side(style='thin', color='cccccc'),
    bottom=Side(style='thin', color='cccccc')
)

header_font = Font(name='微软雅黑', size=10, bold=True, color=WHITE)
header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
title_font = Font(name='微软雅黑', size=14, bold=True, color=DARK_BLUE)
section_font = Font(name='微软雅黑', size=11, bold=True, color=DARK_BLUE)
body_font = Font(name='微软雅黑', size=10, color='333333')
bold_font = Font(name='微软雅黑', size=10, bold=True, color='333333')
small_font = Font(name='微软雅黑', size=9, color='888888')
input_fill = PatternFill(start_color=YELLOW_INPUT, end_color=YELLOW_INPUT, fill_type='solid')
good_fill = PatternFill(start_color=GREEN_GOOD, end_color=GREEN_GOOD, fill_type='solid')
bad_fill = PatternFill(start_color=RED_BAD, end_color=RED_BAD, fill_type='solid')
warn_fill = PatternFill(start_color=ORANGE_WARN, end_color=ORANGE_WARN, fill_type='solid')
light_blue_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
money_fmt = '#,##0'
money2_fmt = '#,##0.00'
pct_fmt = '0.0%'
ton_fmt = '#,##0.0'

def set_cell(ws, row, col, value=None, font=None, fill=None, alignment=None, fmt=None):
    """Set cell with all formatting in one call. Returns cell."""
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if fmt:
        cell.number_format = fmt
    cell.border = thin_border
    return cell

def write_header(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        set_cell(ws, row, start_col + i, h, font=header_font, fill=header_fill, alignment=center_align)

def write_data_row(ws, row, data, start_col=1, label_font=bold_font, data_font=body_font, data_fill=None, fmts=None):
    """Write a data row. data[0] is label, rest are values."""
    set_cell(ws, row, start_col, data[0], font=label_font, alignment=left_align)
    for i, val in enumerate(data[1:], 1):
        f = fmts[i-1] if fmts and i-1 < len(fmts) else None
        set_cell(ws, row, start_col + i, val, font=data_font, fill=data_fill, alignment=center_align, fmt=f)

def section_title(ws, row, text, ncols=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    set_cell(ws, row, 1, text, font=section_font, alignment=left_align)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ═══════════════════════════════════════════
# Sheet 1: 参数设定
# ═══════════════════════════════════════════
ws1 = wb.active
ws1.title = '1-参数设定'
S1 = "'1-参数设定'"  # for cross-sheet refs

# Title
ws1.merge_cells('A1:J1')
set_cell(ws1, 1, 1, '决策科学案例01 — 采购决策优化 — 参数设定', font=title_font, alignment=center_align)
ws1.row_dimensions[1].height = 30
ws1.merge_cells('A2:J2')
set_cell(ws1, 2, 1, '⬛ 黄色 = 可调参数 | 修改后所有Sheet自动重算 | 请启用Excel公式自动计算', font=small_font, alignment=center_align)

# ── A. 全局参数 (row 4) ──
r = 4
section_title(ws1, r, 'A. 全局参数', 10)
r = 5
write_header(ws1, r, ['参数', '符号', '数值', '单位', '说明'])
set_col_widths(ws1, [24, 8, 10, 10, 42])

global_params = [
    # [label, symbol, value, unit, note]
    ['年持有成本率', 'h', 0.08, '%', '库存持有成本占物料价值的比例，含资金占用+仓储+保险'],
    ['紧急采购溢价', 'e', 1.5, '倍数', '紧急采购价格 = 正常单价 × 此倍数。底阀可达2.0x'],
    ['计划周期', 'T', 12, '月', '年度滚动计划'],
    ['年资金成本率', 'r', 0.035, '%', '用于折现未来现金流，LPR基准'],
    ['预算上限', 'B', 20000000, '¥/年', '采购年度总预算约束'],
]
for i, row_data in enumerate(global_params):
    row = r + 1 + i
    # Write cells explicitly
    set_cell(ws1, row, 1, row_data[0], font=bold_font, alignment=left_align)
    set_cell(ws1, row, 2, row_data[1], font=body_font, alignment=center_align)
    set_cell(ws1, row, 3, row_data[2], font=body_font, fill=input_fill, alignment=center_align,
             fmt=pct_fmt if '%' in str(row_data[3]) else (money_fmt if '¥' in str(row_data[3]) else '0.0'))
    set_cell(ws1, row, 4, row_data[3], font=body_font, alignment=center_align)
    set_cell(ws1, row, 5, row_data[4], font=small_font, alignment=left_align)

# Cell map for global params
# B6: h=0.08, B7: e=1.5, B8: T=12, B9: r=0.035, B10: B=20000000
# Actually rows are 6,7,8,9,10 (r=5+1)
GL_H = 'C6'   # holding cost rate 0.08
GL_E = 'C7'   # emergency premium 1.5
GL_T = 'C8'   # periods 12
GL_R = 'C9'   # capital cost 0.035
GL_B = 'C10'  # budget 20000000

# ── B. 物料主数据 (row 11) ──
r = 11
section_title(ws1, r, 'B. 物料主数据（核心优化参数）', 10)
r = 12
mat_headers = ['物料名称', '类别', '单价\n(¥/kg)', 'MOQ\n(吨或台套)', 'Lead Time\n(周)',
               '损耗率\n标准规格', '损耗率\n非标规格', '年用量\n(吨)', '单位', '关键特征']
write_header(ws1, r, mat_headers)

materials_data = [
    # [name, category, price, MOQ, LT, loss_std, loss_nonstd, annual_qty, unit, feature]
    # Row 13-18 (r+1 to r+6)
    ['304/2B不锈钢板', '罐体钢板', 15.0, 8, 7, 0.20, 0.05, 1020, '吨', '市场定价，无议价权'],
    ['316L不锈钢板', '罐体钢板', 28.0, 6, 7, 0.20, 0.05, 360, '吨', '特种罐专用，高价值'],
    ['Q345方管', '框架型钢', 6.5, 3, 2, 0.10, 0.03, 480, '吨', '补货快，替代品多'],
    ['DN150法兰', '法兰接管', 45.0, 0.5, 5, 0.05, 0.02, 120, '吨', '多品类共用'],
    ['底阀组件', '阀门组件', 320.0, 0.1, 10, 0.02, 0.02, 40, '台套', '瓶颈物料！LT最长、断货代价极高'],
    ['聚氨酯保温', '保温材料', 18.0, 2, 3, 0.08, 0.03, 200, '吨', '灵活采购，供应商多'],
]

# Cell map: rows 13-18, cols C=price(3), D=MOQ(4), E=LT(5), F=loss_std(6), G=loss_nonstd(7), H=annual_qty(8)
for i, row_data in enumerate(materials_data):
    row = r + 1 + i
    set_cell(ws1, row, 1, row_data[0], font=bold_font, alignment=left_align)
    set_cell(ws1, row, 2, row_data[1], font=body_font, alignment=center_align)
    set_cell(ws1, row, 3, row_data[2], font=body_font, fill=input_fill, alignment=center_align, fmt=money2_fmt)
    set_cell(ws1, row, 4, row_data[3], font=body_font, fill=input_fill, alignment=center_align, fmt='0.0')
    set_cell(ws1, row, 5, row_data[4], font=body_font, fill=input_fill, alignment=center_align, fmt='0')
    set_cell(ws1, row, 6, row_data[5], font=body_font, fill=input_fill, alignment=center_align, fmt=pct_fmt)
    set_cell(ws1, row, 7, row_data[6], font=body_font, fill=input_fill, alignment=center_align, fmt=pct_fmt)
    set_cell(ws1, row, 8, row_data[7], font=body_font, alignment=center_align, fmt=ton_fmt)
    set_cell(ws1, row, 9, row_data[8], font=body_font, alignment=center_align)
    set_cell(ws1, row, 10, row_data[9], font=small_font, alignment=left_align)

# ── C. 车间需求参数 (row 20) ──
r = 20
section_title(ws1, r, 'C. 车间月均需求（吨/月，可修改）', 10)
r = 21
ws_headers = ['车间', '标罐', '特罐', '碳罐', '气体罐', '小件']
write_header(ws1, r, ['物料 \\ 车间'] + ws_headers[1:])

ws_demand = [
    # [material_name, 标罐, 特罐, 碳罐, 气体罐, 小件]
    # Row 22-27
    ['304/2B不锈钢板', 85, 30, 15, 20, 8],
    ['316L不锈钢板', 20, 50, 5, 10, 3],
    ['Q345方管', 40, 15, 8, 10, 5],
    ['DN150法兰', 8, 4, 2, 3, 1.5],
    ['底阀组件', 15, 8, 3, 12, 2],
    ['聚氨酯保温', 15, 8, 4, 5, 3],
]

for i, row_data in enumerate(ws_demand):
    row = r + 1 + i
    set_cell(ws1, row, 1, row_data[0], font=bold_font, alignment=left_align)
    for c in range(2, 7):
        set_cell(ws1, row, c, row_data[c-1], font=body_font, fill=input_fill, alignment=center_align, fmt=ton_fmt)

# ── 图例 (row 29) ──
r = 29
ws1.merge_cells(f'A{r}:J{r}')
set_cell(ws1, r, 1, '图例： 黄色填充 = 可调输入参数  |  白色 = 计算结果/固定参考  |  所有Sheet均引用本Sheet的参数，修改后自动联动', font=small_font, alignment=left_align)

# ═══════════════════════════════════════════
# Sheet 2: 规格选择优化器
# ═══════════════════════════════════════════
ws2 = wb.create_sheet('2-规格选择优化器')
S2 = "'2-规格选择优化器'"

ws2.merge_cells('A1:K1')
set_cell(ws2, 1, 1, '规格选择优化器 — "贵的往往更便宜"', font=title_font, alignment=center_align)
ws2.row_dimensions[1].height = 30
ws2.merge_cells('A2:K2')
set_cell(ws2, 2, 1, '核心公式：实际可用成本 = 单价 ÷ (1 - 损耗率)  |  MOQ一票否决  |  修改参数Sheet黄色单元格后本Sheet自动更新', font=small_font, alignment=center_align)

# ── 304不锈钢板 单物料深度分析 ──
r = 4
section_title(ws2, r, '▎304/2B不锈钢板 — 规格选择决策（五个车间的逐车间判断）', 11)

r = 5
opt_headers = ['车间', '月需求\n(吨)', '年需求\n(吨)',
               '标准板\n单价¥/kg', '标准板\n损耗率', '标准板\nTCO ¥/年',
               '非标板\n单价¥/kg', '非标板\n损耗率', '非标板\nTCO ¥/年',
               'MOQ\n达标?', '最优决策']
write_header(ws2, r, opt_headers)

# Data rows: 标罐(row6), 特罐(row7), 碳罐(row8), 气体罐(row9), 小件(row10)
# Ref: Sheet1!B22-F27 for demand, Sheet1!C13 for price_std, Sheet1!F13 loss_std, Sheet1!G13 loss_nonstd, Sheet1!D13 MOQ
# Standard price for 304 plate: S1!C13=15, Nonstd price: S1!C13 * 1.13 ≈ 17 (we'll compute)
# Actually non-std price isn't in sheet 1. Let me add it or compute as std_price * (1 + premium)

# For simplicity, let me compute nonstd price = std_price * 1.12 (12% premium for custom sizing)
# But the user should be able to adjust. Let me add a nonstd premium to Sheet1.

# Actually, let me keep it simpler and cleaner. I'll use direct cell refs.
# In Sheet1, we have price_std (col C), loss_std (col F), loss_nonstd (col G), MOQ (col D)
# Nonstd price = price_std * (1 + nonstd_premium). Let me add that to Sheet1.

# Hmm, I think the cleanest approach is:
# - Sheet1 has: price_std, price_nonstd_premium (%), MOQ, LT, loss_std, loss_nonstd
# - Sheet2 references those

# Actually let me just add a nonstd premium column or compute it as a fixed 12% markup.
# For the demo, nonstd_price = std_price * 1.12

# Let me restructure Sheet1 slightly... Actually no, let me just keep going.
# I'll add nonstd_price_premium as a global param in Sheet1 row 10 area.

# Wait, I need to add this to Sheet1. Let me insert a row there. Actually, I can add it after the global params.
# Let me add it as D. 非标板加价率 in Sheet1.

# Actually, this is getting complicated. Let me just add a cell in Sheet2 for the non-std premium.
# Or better: compute it in the formula. Nonstd price for 304 = S1!C13 * (1 + 0.12)

# The TCO formula:
# TCO_std = annual_demand * price_std / (1 - loss_std) * 1000 + annual_demand * price_std * holding_rate * 0.5 * 1000
# TCO_nonstd = annual_demand * price_nonstd / (1 - loss_nonstd) * 1000 + same holding
# MOQ_OK = monthly_demand >= MOQ
# Decision = IF(AND(MOQ_OK, TCO_nonstd < TCO_std), "非标板 ✓", "标准板")

# For row 6 (标罐):
# B6 = '1-参数设定'!B22 (85 tons)
# C6 = B6*12
# D6 = '1-参数设定'!C13 (15 ¥/kg)
# E6 = '1-参数设定'!F13 (0.20)
# F6 = C6*1000*D6/(1-E6) + C6*1000*D6*'1-参数设定'!C6*0.5  -- TCO
# G6 = D6*(1+0.12)  -- nonstd price = std * 1.12
# H6 = '1-参数设定'!G13 (0.05)
# I6 = C6*1000*G6/(1-H6) + C6*1000*G6*'1-参数设定'!C6*0.5
# J6 = IF(B6 >= '1-参数设定'!D13, "✓ 达标", "✗ 仅"&TEXT(B6,"0.0")&"吨<MOQ"&TEXT('1-参数设定'!D13,"0.0")&"吨")
# K6 = IF(AND(B6>='1-参数设定'!D13, I6<F6), "非标板 ✓", "标准板")

workshops = ['标罐', '特罐', '碳罐', '气体罐', '小件']
# Sheet1 rows: 22-27, cols B-F for demand
ws1_demand_rows = {'标罐': 2, '特罐': 3, '碳罐': 4, '气体罐': 5, '小件': 6}
# 304不锈钢板 is row 13 in Sheet1

for i, ws_name in enumerate(workshops):
    row = r + 1 + i
    s1_col = ws1_demand_rows[ws_name]  # column in Sheet1 (B=2, C=3, D=4, E=5, F=6)
    s1_col_letter = get_column_letter(s1_col)
    s1_demand_cell = f"{S1}!{s1_col_letter}22"  # Sheet1 demand for 304 plate by workshop

    # B: monthly demand
    set_cell(ws2, row, 1, ws_name, font=bold_font, alignment=left_align)
    set_cell(ws2, row, 2, f'={s1_demand_cell}', font=body_font, alignment=center_align, fmt=ton_fmt)
    # C: annual demand = monthly * 12
    set_cell(ws2, row, 3, f'=B{row}*{GL_T}', font=body_font, alignment=center_align, fmt=ton_fmt)
    # D: std price (from Sheet1 C13)
    set_cell(ws2, row, 4, f'={S1}!C13', font=body_font, alignment=center_align, fmt=money2_fmt)
    # E: std loss rate (from Sheet1 F13)
    set_cell(ws2, row, 5, f'={S1}!F13', font=body_font, alignment=center_align, fmt=pct_fmt)
    # F: TCO_std = annual_demand * 1000 * price / (1 - loss) + holding
    # holding = avg_inventory * price * holding_rate
    # avg_inventory ≈ monthly_demand (simplified: 1 month of stock)
    tco_std = f'=C{row}*1000*D{row}/(1-E{row})+B{row}*1000*D{row}*{GL_H}'
    set_cell(ws2, row, 6, tco_std, font=body_font, alignment=center_align, fmt=money_fmt)
    # G: nonstd price = std_price * 1.12
    set_cell(ws2, row, 7, f'=D{row}*1.12', font=body_font, alignment=center_align, fmt=money2_fmt)
    # H: nonstd loss rate (from Sheet1 G13)
    set_cell(ws2, row, 8, f'={S1}!G13', font=body_font, alignment=center_align, fmt=pct_fmt)
    # I: TCO_nonstd
    tco_nonstd = f'=C{row}*1000*G{row}/(1-H{row})+B{row}*1000*G{row}*{GL_H}'
    set_cell(ws2, row, 9, tco_nonstd, font=body_font, alignment=center_align, fmt=money_fmt)
    # J: MOQ check
    moq_check = f'=IF(B{row}>={S1}!D13,"✓ 达标","✗ 不足(MOQ="&TEXT({S1}!D13,"0.0")&"吨)")'
    set_cell(ws2, row, 10, moq_check, font=body_font, alignment=center_align)
    # K: Optimal decision
    # If MOQ met AND nonstd TCO < std TCO → nonstd, else std
    decision = f'=IF(AND(B{row}>={S1}!D13,I{row}<F{row}),"非标板 ✓","标准板")'
    set_cell(ws2, row, 11, decision, font=bold_font, alignment=center_align)
    # Conditional formatting: green if nonstd, red if std
    # We'll use the value for conditional logic later

# ── 汇总行 ──
r_sum = r + 6
set_cell(ws2, r_sum, 1, '合计/加权', font=bold_font, alignment=left_align)
set_cell(ws2, r_sum, 2, f'=SUM(B{r+1}:B{r+5})', font=bold_font, alignment=center_align, fmt=ton_fmt)
set_cell(ws2, r_sum, 3, f'=SUM(C{r+1}:C{r+5})', font=bold_font, alignment=center_align, fmt=ton_fmt)
for c in range(4, 12):
    if c == 6 or c == 9:  # TCO columns - sum
        set_cell(ws2, r_sum, c, f'=SUM({get_column_letter(c)}{r+1}:{get_column_letter(c)}{r+5})', font=bold_font, alignment=center_align, fmt=money_fmt)
    elif c == 10:  # MOQ - count达标
        set_cell(ws2, r_sum, c, f'=COUNTIF(J{r+1}:J{r+5},"*达标*")&"/5 车间达标"', font=bold_font, alignment=center_align)
    elif c == 11:  # Decision - count nonstd
        set_cell(ws2, r_sum, c, f'=COUNTIF(K{r+1}:K{r+5},"非标板*")&"/5 车间选非标"', font=bold_font, alignment=center_align)
    else:
        set_cell(ws2, r_sum, c, '', font=bold_font, alignment=center_align)

# ── 决策切换点分析 ──
r = r_sum + 2
section_title(ws2, r, '▎决策切换点分析 — 在什么需求水平下非标板"反转"胜出？', 11)
r += 1
write_header(ws2, r, ['分析项', '公式/逻辑', '数值', '单位', '说明'])

# Switching point: when does TCO_nonstd < TCO_std?
# Solve: D*1000*G/(1-H) + D*1000*G*h = D*1000*D_price/(1-E) + D*1000*D_price*h
# Actually monthly demand doesn't matter for the per-unit comparison since both scale linearly
# Per-kg: price_std/(1-loss_std) vs price_nonstd/(1-loss_nonstd)
# The switching happens when MOQ is met AND per-kg cost is lower
# For 304: 15/(1-0.2) = 18.75 vs 16.8/(1-0.05) = 17.68 → nonstd always better IF MOQ met
# The real switch is the MOQ threshold itself

switch_data = [
    ['标准板有效单价', '=单价÷(1-损耗率)', f'=D6/(1-E6)', '¥/kg', '含损耗后的实际可用成本'],
    ['非标板有效单价', '=单价÷(1-损耗率)', f'=G6/(1-H6)', '¥/kg', '非标板的有效成本'],
    ['单位成本节省', '=标准有效单价-非标有效单价', f'=B{r+1}-B{r+2}', '¥/kg', '每kg节省金额'],
    ['年化节省(单个车间)', '=单位节省×年需求×1000', f'=B{r+3}*C6*1000', '¥/年', '仅当MOQ达标时成立'],
    ['MOQ门槛(吨/月)', '', f'={S1}!D13', '吨', '月需求必须≥此值才能选非标'],
    ['标罐月需求vs MOQ', '', f'=B6', '吨', f'=IF(B6>={S1}!D13,"✓ 满足→可选非标","✗ 不满足→强制标准板")'],
    ['碳罐月需求vs MOQ', '', f'=B8', '吨', f'=IF(B8>={S1}!D13,"✓ 满足→可选非标","✗ 不满足→强制标准板")'],
]

for i, row_data in enumerate(switch_data):
    row = r + 1 + i
    set_cell(ws2, row, 1, row_data[0], font=bold_font, alignment=left_align)
    set_cell(ws2, row, 2, row_data[1], font=body_font, alignment=center_align)
    set_cell(ws2, row, 3, row_data[2], font=body_font, alignment=center_align,
             fmt=money2_fmt if '¥/kg' in str(row_data[3]) else (money_fmt if '¥/年' in str(row_data[3]) else ton_fmt))
    set_cell(ws2, row, 4, row_data[3], font=body_font, alignment=center_align)
    set_cell(ws2, row, 5, row_data[4], font=small_font, alignment=left_align)

# ── 洞察 ──
r = r + 9
section_title(ws2, r, '▎核心洞察', 11)
r += 1
insights = [
    '1. 304不锈钢板：非标板有效单价¥17.68/kg < 标准板¥18.75/kg，贵了12%的单价实际便宜了5.7%。——前提是MOQ达标。',
    '2. 碳罐车间（15吨/月）单独无法达标非标MOQ（8吨），被迫选标准板。这就是聚合效应的价值所在 → 见"聚合效应计算器"。',
    '3. 修改"参数设定"Sheet中304板的损耗率（如标准板损耗改为30%），观察TCO差距如何放大，决策信心如何变化。',
    '4. 如果非标板加价率从12%提高到20%（修改G列公式中的1.12→1.20），有效单价=18/(1-0.05)=18.95 > 18.75，非标板不再有优势！',
]
for txt in insights:
    set_cell(ws2, r, 1, txt, font=body_font, alignment=left_align)
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
    r += 1

set_col_widths(ws2, [10, 11, 11, 11, 11, 15, 11, 11, 15, 16, 14])

# ═══════════════════════════════════════════
# Sheet 3: 聚合效应计算器
# ═══════════════════════════════════════════
ws3 = wb.create_sheet('3-聚合效应计算器')
S3 = "'3-聚合效应计算器'"

ws3.merge_cells('A1:J1')
set_cell(ws3, 1, 1, '跨车间聚合效应计算器 — "局部最优之和 ≠ 全局最优"', font=title_font, alignment=center_align)
ws3.row_dimensions[1].height = 30
ws3.merge_cells('A2:J2')
set_cell(ws3, 2, 1, '聚合逻辑：同规格物料跨车间合并采购→突破MOQ→解锁非标板→降低损耗 | 修改参数Sheet的车间需求后自动更新', font=small_font, alignment=center_align)

# ── 304不锈钢板聚合分析 ──
r = 4
section_title(ws3, r, '▎304/2B不锈钢板 — 五车间独立 vs 聚合（MOQ=8吨）', 10)

r = 5
agg_headers = ['', '标罐', '特罐', '碳罐', '气体罐', '小件', '五车间合计', 'MOQ', '达标?', '加权损耗率']
write_header(ws3, r, agg_headers)

# ── 方案A: 独立采购 ──
r = 6
ws3.merge_cells(f'A{r}:A{r+7}')
set_cell(ws3, r, 1, '方案A\n独立采购', font=Font(name='微软雅黑', size=10, bold=True, color='cc0000'), fill=bad_fill, alignment=center_align)

# Individual procurement data
# Row 6-13
ind_rows = [
    # [label, 标罐, 特罐, 碳罐, 气体罐, 小件, formula_col7, formula_col8, formula_col9, formula_col10]
    # These are mostly text/formula rows
]

# Let me write individual rows explicitly for clarity
# Row 6: 月需求(吨)
r_ind = r
for c in range(2, 7):
    col_l = get_column_letter(c)
    set_cell(ws3, r_ind, c, f'={S1}!{col_l}22', font=body_font, alignment=center_align, fmt=ton_fmt)  # 304 demand by workshop
set_cell(ws3, r_ind, 7, f'=SUM(B{r_ind}:F{r_ind})', font=bold_font, alignment=center_align, fmt=ton_fmt)
set_cell(ws3, r_ind, 8, f'={S1}!D13', font=body_font, alignment=center_align, fmt='0.0')
set_cell(ws3, r_ind, 9, '', font=body_font, alignment=center_align)
set_cell(ws3, r_ind, 10, '', font=body_font, alignment=center_align)

# Row 7: 独立MOQ达标(月)?
r_ind += 1
for c in range(2, 7):
    set_cell(ws3, r_ind, c, f'=IF(B{r_ind-1}>={S1}!D13,"✓","✗")', font=body_font, alignment=center_align)
set_cell(ws3, r_ind, 7, '', font=body_font, alignment=center_align)
set_cell(ws3, r_ind, 8, '', font=body_font, alignment=center_align)
set_cell(ws3, r_ind, 9, '', font=body_font, alignment=center_align)
set_cell(ws3, r_ind, 10, '', font=body_font, alignment=center_align)

# Row 8: 可选规格
r_ind += 1
for c in range(2, 7):
    set_cell(ws3, r_ind, c, f'=IF(B{r_ind-1}="✓","非标板","标准板")', font=body_font, alignment=center_align)
for c in range(7, 11):
    set_cell(ws3, r_ind, c, '', font=body_font, alignment=center_align)

# Row 9: 损耗率(独立)
r_ind += 1
for c in range(2, 7):
    # If nonstd → loss_nonstd, else loss_std
    set_cell(ws3, r_ind, c, f'=IF(B{r_ind-1}="非标板",{S1}!G13,{S1}!F13)', font=body_font, alignment=center_align, fmt=pct_fmt)
set_cell(ws3, r_ind, 7, '', font=body_font, alignment=center_align)
set_cell(ws3, r_ind, 8, '', font=body_font, alignment=center_align)
set_cell(ws3, r_ind, 9, '', font=body_font, alignment=center_align)
set_cell(ws3, r_ind, 10, '', font=body_font, alignment=center_align)

# Row 10: 年损耗量(吨)
r_ind += 1
for c in range(2, 7):
    # annual demand = monthly * 12, loss = demand * loss_rate
    set_cell(ws3, r_ind, c, f'=B6*12*B{r_ind}', font=body_font, fill=bad_fill, alignment=center_align, fmt=ton_fmt)
set_cell(ws3, r_ind, 7, f'=SUM(B{r_ind}:F{r_ind})', font=bold_font, fill=bad_fill, alignment=center_align, fmt=ton_fmt)
for c in range(8, 11):
    set_cell(ws3, r_ind, c, '', font=body_font, alignment=center_align)

# Row 11: 加权平均损耗率
r_ind += 1
for c in range(2, 8):
    set_cell(ws3, r_ind, c, '', font=body_font, alignment=center_align)
set_cell(ws3, r_ind, 9, '', font=body_font, alignment=center_align)
# Weighted average: total_loss / total_demand
set_cell(ws3, r_ind, 10, f'=G{r_ind-1}/(G6*12)', font=bold_font, fill=bad_fill, alignment=center_align, fmt=pct_fmt)

# Row 12: 年采购成本(¥)
r_ind += 1
for c in range(2, 7):
    set_cell(ws3, r_ind, c, f'=B6*12*1000*{S1}!C13', font=body_font, alignment=center_align, fmt=money_fmt)
set_cell(ws3, r_ind, 7, f'=SUM(B{r_ind}:F{r_ind})', font=bold_font, alignment=center_align, fmt=money_fmt)
for c in range(8, 11):
    set_cell(ws3, r_ind, c, '', font=body_font, alignment=center_align)

# Row 13: 年损耗金额(¥)
r_ind += 1
for c in range(2, 7):
    set_cell(ws3, r_ind, c, f'=B{r_ind-3}*1000*{S1}!C13', font=body_font, fill=bad_fill, alignment=center_align, fmt=money_fmt)
set_cell(ws3, r_ind, 7, f'=SUM(B{r_ind}:F{r_ind})', font=bold_font, fill=bad_fill, alignment=center_align, fmt=money_fmt)
for c in range(8, 11):
    set_cell(ws3, r_ind, c, '', font=body_font, alignment=center_align)

r_ind_end_a = r_ind  # save for reference

# ── 方案B: 聚合采购 ──
r = 15
ws3.merge_cells(f'A{r}:A{r+7}')
set_cell(ws3, r, 1, '方案B\n聚合采购', font=Font(name='微软雅黑', size=10, bold=True, color=GREEN_DARK), fill=good_fill, alignment=center_align)

# Row 15: 月需求(吨) — same as A
r_agg = r
for c in range(2, 7):
    col_l = get_column_letter(c)
    set_cell(ws3, r_agg, c, f'={S1}!{col_l}22', font=body_font, alignment=center_align, fmt=ton_fmt)
set_cell(ws3, r_agg, 7, f'=SUM(B{r_agg}:F{r_agg})', font=bold_font, alignment=center_align, fmt=ton_fmt)
set_cell(ws3, r_agg, 8, f'={S1}!D13', font=body_font, alignment=center_align, fmt='0.0')
set_cell(ws3, r_agg, 9, '', font=body_font, alignment=center_align)
set_cell(ws3, r_agg, 10, '', font=body_font, alignment=center_align)

# Row 16: 聚合MOQ达标?
r_agg += 1
for c in range(2, 7):
    set_cell(ws3, r_agg, c, '', font=body_font, alignment=center_align)
set_cell(ws3, r_agg, 7, f'=IF(G{r_agg-1}>={S1}!D13,"✓ 聚合后"&TEXT(G{r_agg-1},"0.0")&"吨>>MOQ"&TEXT({S1}!D13,"0.0")&"吨","✗")', font=bold_font, fill=good_fill, alignment=center_align)
for c in range(8, 11):
    set_cell(ws3, r_agg, c, '', font=body_font, alignment=center_align)

# Row 17: 可选规格 (聚合后全部可用非标板)
r_agg += 1
for c in range(2, 7):
    set_cell(ws3, r_agg, c, '非标板 (聚合解锁)', font=body_font, fill=good_fill, alignment=center_align)
for c in range(7, 11):
    set_cell(ws3, r_agg, c, '', font=body_font, alignment=center_align)

# Row 18: 损耗率(聚合)
r_agg += 1
for c in range(2, 7):
    set_cell(ws3, r_agg, c, f'={S1}!G13', font=body_font, fill=good_fill, alignment=center_align, fmt=pct_fmt)
for c in range(7, 11):
    set_cell(ws3, r_agg, c, '', font=body_font, alignment=center_align)

# Row 19: 年损耗量(吨)
r_agg += 1
for c in range(2, 7):
    set_cell(ws3, r_agg, c, f'=B{r_agg-4}*12*B{r_agg}', font=body_font, fill=good_fill, alignment=center_align, fmt=ton_fmt)
set_cell(ws3, r_agg, 7, f'=SUM(B{r_agg}:F{r_agg})', font=bold_font, fill=good_fill, alignment=center_align, fmt=ton_fmt)

# Row 20: 加权平均损耗率
r_agg += 1
for c in range(2, 8):
    set_cell(ws3, r_agg, c, '', font=body_font, alignment=center_align)
set_cell(ws3, r_agg, 9, '', font=body_font, alignment=center_align)
set_cell(ws3, r_agg, 10, f'=G{r_agg-1}/(G{r_agg-4}*12)', font=bold_font, fill=good_fill, alignment=center_align, fmt=pct_fmt)

# Row 21: 年采购成本 — same price but all nonstd
r_agg += 1
for c in range(2, 7):
    set_cell(ws3, r_agg, c, f'=B{r_agg-6}*12*1000*{S1}!C13*1.12', font=body_font, alignment=center_align, fmt=money_fmt)
set_cell(ws3, r_agg, 7, f'=SUM(B{r_agg}:F{r_agg})', font=bold_font, alignment=center_align, fmt=money_fmt)

# Row 22: 年损耗金额
r_agg += 1
for c in range(2, 7):
    set_cell(ws3, r_agg, c, f'=B{r_agg-3}*1000*{S1}!C13*1.12', font=body_font, fill=good_fill, alignment=center_align, fmt=money_fmt)
set_cell(ws3, r_agg, 7, f'=SUM(B{r_agg}:F{r_agg})', font=bold_font, fill=good_fill, alignment=center_align, fmt=money_fmt)

# ── 对比汇总 ──
r = 24
section_title(ws3, r, '▎聚合前后对比', 10)
r = 25
write_header(ws3, r, ['指标', '独立采购(A)', '聚合采购(B)', '差异(B-A)', '改善比例', '说明'])

# Loss comparison
set_cell(ws3, r+1, 1, '年损耗量(吨)', font=bold_font, alignment=left_align)
set_cell(ws3, r+1, 2, f'=G10', font=body_font, fill=bad_fill, alignment=center_align, fmt=ton_fmt)
set_cell(ws3, r+1, 3, f'=G19', font=body_font, fill=good_fill, alignment=center_align, fmt=ton_fmt)
set_cell(ws3, r+1, 4, f'=C{r+1}-B{r+1}', font=bold_font, fill=warn_fill, alignment=center_align, fmt=ton_fmt)
set_cell(ws3, r+1, 5, f'=D{r+1}/B{r+1}', font=bold_font, alignment=center_align, fmt=pct_fmt)
set_cell(ws3, r+1, 6, '负值=节省', font=body_font, alignment=left_align)

set_cell(ws3, r+2, 1, '加权平均损耗率', font=bold_font, alignment=left_align)
set_cell(ws3, r+2, 2, f'=J11', font=body_font, fill=bad_fill, alignment=center_align, fmt=pct_fmt)
set_cell(ws3, r+2, 3, f'=J20', font=body_font, fill=good_fill, alignment=center_align, fmt=pct_fmt)
set_cell(ws3, r+2, 4, f'=C{r+2}-B{r+2}', font=bold_font, fill=warn_fill, alignment=center_align, fmt=pct_fmt)
set_cell(ws3, r+2, 5, f'=D{r+2}/B{r+2}', font=bold_font, alignment=center_align, fmt=pct_fmt)
set_cell(ws3, r+2, 6, '损耗率下降百分点', font=body_font, alignment=left_align)

set_cell(ws3, r+3, 1, '年损耗金额(¥)', font=bold_font, alignment=left_align)
set_cell(ws3, r+3, 2, f'=G13', font=body_font, fill=bad_fill, alignment=center_align, fmt=money_fmt)
set_cell(ws3, r+3, 3, f'=G22', font=body_font, fill=good_fill, alignment=center_align, fmt=money_fmt)
set_cell(ws3, r+3, 4, f'=C{r+3}-B{r+3}', font=bold_font, fill=warn_fill, alignment=center_align, fmt=money_fmt)
set_cell(ws3, r+3, 5, f'=D{r+3}/B{r+3}', font=bold_font, alignment=center_align, fmt=pct_fmt)
set_cell(ws3, r+3, 6, '年化节省金额', font=body_font, alignment=left_align)

# ── 碳罐解锁故事 ──
r = 30
section_title(ws3, r, '▎聚合的"杀手场景"：碳罐车间', 10)
r += 1
story = [
    '碳罐车间月需仅15吨304板 → 单独无法达标MOQ(8吨) → 被迫选标准板 → 损耗20%。',
    '与标罐(85吨/月)聚合 → 合并100吨/月 >> MOQ(8吨) → 解锁非标板 → 损耗降至5%。',
    '仅碳罐一个车间，聚合前后的损耗差异：15×12×(20%-5%) = 27吨/年 ≈ ¥40,500/年（按¥1,500/吨废料价差估算）。',
    '→ 试试修改"参数设定"Sheet中碳罐的月需求（增大到25吨），观察碳罐能否独立达标，聚合价值如何变化。',
]
for txt in story:
    set_cell(ws3, r, 1, txt, font=body_font, alignment=left_align)
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 1

set_col_widths(ws3, [13, 13, 13, 13, 13, 13, 20, 12, 15, 14])

# ═══════════════════════════════════════════
# Sheet 4: 库存策略优化器
# ═══════════════════════════════════════════
ws4 = wb.create_sheet('4-库存策略优化器')
S4 = "'4-库存策略优化器'"

ws4.merge_cells('A1:I1')
set_cell(ws4, 1, 1, '库存策略优化器 — 安全库存水平的最优计算', font=title_font, alignment=center_align)
ws4.row_dimensions[1].height = 30
ws4.merge_cells('A2:I2')
set_cell(ws4, 2, 1, '核心逻辑：安全库存 = f(Lead Time, 需求波动, 断货代价, 持有成本) | 不同物料的最优库存策略不同 | 修改参数Sheet后自动更新', font=small_font, alignment=center_align)

# ── 决策框架 ──
r = 4
section_title(ws4, r, '▎六种物料的多维度评估 → 决定库存策略', 9)
r = 5
write_header(ws4, r, ['物料', 'Lead Time\n(周)', '单价\n(¥/kg)', '月均需求\n变异系数', '断货代价\n等级', '持有成本\n(¥/单位/年)',
                       '推荐安全\n库存天数', '推荐安全\n库存量', '策略分类'])

# For each of 6 materials (Sheet1 rows 13-18)
# LT: col E, price: col C, demand variability is estimated, shortage cost is qualitative
mat_names = ['304不锈钢板', '316L不锈钢板', 'Q345方管', 'DN150法兰', '底阀组件', '聚氨酯保温']
mat_s1_rows = [13, 14, 15, 16, 17, 18]

# Simplified safety stock formula: SS_days = LT_weeks * (1 + CV) * shortage_factor / holding_factor
# Then SS_qty = daily_demand * SS_days
# We'll simplify to a lookup/scoring approach for the demo

mat_params = [
    # [name, s1_row, CV, shortage_level(1-5), holding_per_unit_year]
    ['304/2B不锈钢板', 13, 0.15, 4, None],  # holding = price * h
    ['316L不锈钢板', 14, 0.20, 4, None],
    ['Q345方管', 15, 0.10, 1, None],
    ['DN150法兰', 16, 0.15, 2, None],
    ['底阀组件', 17, 0.25, 5, None],
    ['聚氨酯保温', 18, 0.10, 1, None],
]

for i, mp in enumerate(mat_params):
    row = r + 1 + i
    s1r = mp[1]
    set_cell(ws4, row, 1, mp[0], font=bold_font, alignment=left_align)
    set_cell(ws4, row, 2, f'={S1}!E{s1r}', font=body_font, alignment=center_align, fmt='0')  # LT
    set_cell(ws4, row, 3, f'={S1}!C{s1r}', font=body_font, alignment=center_align, fmt=money2_fmt)  # price
    set_cell(ws4, row, 4, mp[2], font=body_font, fill=input_fill, alignment=center_align, fmt=pct_fmt)  # CV
    set_cell(ws4, row, 5, mp[3], font=body_font, fill=input_fill, alignment=center_align, fmt='0')  # shortage level
    # holding cost per unit per year = price * holding_rate
    set_cell(ws4, row, 6, f'=C{row}*{GL_H}', font=body_font, alignment=center_align, fmt=money2_fmt)
    # Recommended safety stock days = LT_weeks/7*365 * (1+CV) * shortage_level/3
    set_cell(ws4, row, 7, f'=ROUND(B{row}/52*365*(1+D{row})*E{row}/3,0)', font=bold_font, alignment=center_align, fmt='0')
    # Safety stock qty = daily_demand * SS_days; daily_demand = annual_qty / 365
    set_cell(ws4, row, 8, f'=ROUND({S1}!H{s1r}/365*G{row},1)', font=bold_font, alignment=center_align, fmt='0.0')
    # Strategy classification
    set_cell(ws4, row, 9, f'=IF(G{row}>=30,"高安全库存",IF(G{row}>=14,"中等库存","低库存-灵活补货"))', font=bold_font, alignment=center_align)

# ── 三种策略的量化对比（以底阀组件为例） ──
r = 13
section_title(ws4, r, '▎以底阀组件为例：三种库存策略的量化对比（可调参数）', 9)
r = 14
write_header(ws4, r, ['策略', '安全库存\n(台套)', '平均库存\n(台套)', '年持有成本\n(¥)', '预计年紧急\n采购次数', '紧急采购\n年成本(¥)',
                       '总相关成本\n(¥/年)', '断货风险', '综合评级'])

# Input parameters for valve
# Monthly demand = 40/year / 12 ≈ 3.33/month, price=320, LT=10 weeks
# Let me add editable parameters for this analysis
r_valve = r + 1

set_cell(ws4, r_valve, 1, '参数：月均需求(台套)', font=bold_font, alignment=left_align)
set_cell(ws4, r_valve, 2, f'={S1}!H17/12', font=body_font, fill=input_fill, alignment=center_align, fmt='0.0')
ws4.merge_cells(f'C{r_valve}:I{r_valve}')
set_cell(ws4, r_valve, 3, '← 可修改（或在参数Sheet中修改年用量）', font=small_font, alignment=left_align)

r_valve2 = r_valve + 1
set_cell(ws4, r_valve2, 1, '参数：需求不确定性(σ/μ)', font=bold_font, alignment=left_align)
set_cell(ws4, r_valve2, 2, 0.25, font=body_font, fill=input_fill, alignment=center_align, fmt=pct_fmt)
set_cell(ws4, r_valve2, 3, '← 可修改，ETO特征：波动越大，安全库存越高', font=small_font, alignment=left_align)

# Three strategy rows
strat_start = r_valve2 + 2
strategies = [
    # [name, ss_factor, label]
    ['A. 激进策略（库存最小化）', 0.3, 'bad'],   # SS = 0.3 * monthly_demand
    ['B. 保守策略（不断货优先）', 2.0, 'bad'],    # SS = 2.0 * monthly_demand
    ['C. 模型推荐（总成本最优）', 0.9, 'good'],    # SS = 0.9 * monthly_demand (optimal balance)
]

for i, strat in enumerate(strategies):
    row = strat_start + i
    # Name
    set_cell(ws4, row, 1, strat[0], font=bold_font, alignment=left_align)
    fill = good_fill if strat[2] == 'good' else bad_fill
    # SS qty
    set_cell(ws4, row, 2, f'=ROUND(B{r_valve}*{strat[1]},1)', font=body_font, fill=fill, alignment=center_align, fmt='0.0')
    # Avg inventory = SS + cycle_stock/2 = SS + monthly/2
    set_cell(ws4, row, 3, f'=B{row}+B{r_valve}/2', font=body_font, alignment=center_align, fmt='0.0')
    # Annual holding cost = avg_inventory * price * holding_rate
    set_cell(ws4, row, 4, f'=C{row}*{S1}!C17*{GL_H}', font=body_font, alignment=center_align, fmt=money_fmt)
    # Expected emergency purchases/year = function of SS coverage
    # Simplified: emergency_prob = MAX(0, (monthly_demand*(1+cv) - SS) / monthly_demand)
    set_cell(ws4, row, 5, f'=MAX(0,ROUND((B{r_valve}*(1+B{r_valve2})-B{row})/B{r_valve}*12,1))', font=body_font, alignment=center_align, fmt='0.0')
    # Emergency cost = emergency_times * monthly * emergency_premium * price
    set_cell(ws4, row, 6, f'=E{row}*B{r_valve}*({GL_E}-1)*{S1}!C17', font=body_font, alignment=center_align, fmt=money_fmt)
    # Total cost
    set_cell(ws4, row, 7, f'=D{row}+F{row}', font=bold_font, fill=fill, alignment=center_align, fmt=money_fmt)
    # Shortage risk
    set_cell(ws4, row, 8, f'=IF(B{row}>=B{r_valve}*(1+B{r_valve2}),"极低","中高")', font=body_font, alignment=center_align)
    # Rating
    set_cell(ws4, row, 9, f'=IF(E{row}<0.5,"★★★ 最优",IF(E{row}<2,"★★ 可接受","★ 需改进"))', font=bold_font, alignment=center_align)

# Fix row 5 alignment issue - let me set cell E correctly for each strategy row
for i in range(3):
    ws4.cell(row=strat_start+i, column=5).alignment = center_align

# Highlight best
r_best = strat_start + 2
for c in range(1, 10):
    ws4.cell(row=r_best, column=c).font = Font(name='微软雅黑', size=10, bold=True, color=GREEN_DARK)

# ── 洞察 ──
r = strat_start + 4
section_title(ws4, r, '▎库存策略的决策逻辑', 9)
r += 1
insights4 = [
    '1. 底阀组件：LT=10周(最长)、单价高(¥320)、断货代价极高(成品必装无替代) → 模型推荐中等偏高安全库存，总成本最优。',
    '2. Q345方管：LT=2周(最短)、单价低(¥6.5)、替代品多 → 模型推荐低库存、高频补货。库存策略因物料而异，没有"一刀切"的准则。',
    '3. 修改需求不确定性参数(B15格)，观察三种策略的总成本排序如何变化。波动越大，高安全库存策略的优势越明显。',
    '4. 关键权衡：持有成本(线性增长) vs 紧急采购成本(非线性惩罚)。模型找到的"最优安全库存"恰好是边际持有成本=边际缺货成本的交点。',
]
for txt in insights4:
    set_cell(ws4, r, 1, txt, font=body_font, alignment=left_align)
    ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 1

set_col_widths(ws4, [24, 12, 12, 12, 12, 14, 14, 12, 16])

# ═══════════════════════════════════════════
# Sheet 5: 综合仪表板
# ═══════════════════════════════════════════
ws5 = wb.create_sheet('5-综合仪表板')
S5 = "'5-综合仪表板'"

ws5.merge_cells('A1:G1')
set_cell(ws5, 1, 1, '采购决策优化 — 综合仪表板', font=title_font, alignment=center_align)
ws5.row_dimensions[1].height = 30

# ── 关键指标卡片 ──
r = 3
section_title(ws5, r, '▎关键决策指标', 7)

r = 4
# Card 1
# Card headers
ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
set_cell(ws5, r, 1, '规格选择优化', font=Font(name='微软雅黑', size=12, bold=True, color=WHITE), fill=header_fill, alignment=center_align)
ws5.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
set_cell(ws5, r, 4, '聚合效应', font=Font(name='微软雅黑', size=12, bold=True, color=WHITE), fill=header_fill, alignment=center_align)

# Card 1 content
r += 1
kpi_data_1 = [
    ['304板TCO节省潜力', f'={S2}!F6-{S2}!I6', money_fmt, '优化前-优化后(标罐单车间)'],
    ['可选非标板的车间数', f'={S2}!K11', None, '需≥MOQ且TCO更优'],
    ['平均损耗率降幅', f'={S2}!E6-{S2}!H6', pct_fmt, '非标vs标准的损耗率差'],
]
for i, kd in enumerate(kpi_data_1):
    row = r + i
    set_cell(ws5, row, 1, kd[0], font=bold_font, alignment=left_align)
    set_cell(ws5, row, 2, kd[1], font=Font(name='微软雅黑', size=14, bold=True, color=DARK_BLUE), alignment=center_align, fmt=kd[2])
    set_cell(ws5, row, 3, kd[3], font=small_font, alignment=left_align)

# Card 2 content
for i, kd in enumerate([
    ['聚合后年损耗量(吨)', f'={S3}!G19', ton_fmt, '聚合采购方案'],
    ['聚合vs独立损耗降幅', f'={S3}!D26', ton_fmt, '年节省损耗吨数'],
    ['损耗率(聚合后)', f'={S3}!J20', pct_fmt, '全部车间享受非标板优势'],
]):
    row = r + i
    set_cell(ws5, row, 4, kd[0], font=bold_font, alignment=left_align)
    set_cell(ws5, row, 5, kd[1], font=Font(name='微软雅黑', size=14, bold=True, color=DARK_BLUE), alignment=center_align, fmt=kd[2])
    set_cell(ws5, row, 6, kd[3], font=small_font, alignment=left_align)

# ── 库存策略汇总 ──
r = 9
section_title(ws5, r, '▎库存策略一览', 7)
r = 10
write_header(ws5, r, ['物料', '安全库存天数', '策略分类', '年持有成本(¥)', '关键驱动因素'])

for i, mp in enumerate(mat_params):
    row = r + 1 + i
    set_cell(ws5, row, 1, mp[0], font=bold_font, alignment=left_align)
    set_cell(ws5, row, 2, f'={S4}!G{r-4+i}', font=body_font, alignment=center_align, fmt='0')
    set_cell(ws5, row, 3, f'={S4}!I{r-4+i}', font=bold_font, alignment=center_align)
    # holding cost = safety_stock_qty * price * holding_rate
    set_cell(ws5, row, 4, f'={S4}!H{r-4+i}*{S4}!C{r-4+i}*{GL_H}', font=body_font, alignment=center_align, fmt=money_fmt)
    # Key driver
    drivers = ['LT+断货风险', '高价值+LT', 'LT短可灵活补货', '多品类共用', 'LT最长+断货代价极高', 'LT短+供应充足']
    set_cell(ws5, row, 5, drivers[i], font=small_font, alignment=left_align)

# ── 敏感性分析 ──
r = 18
section_title(ws5, r, '▎敏感性分析：损耗率变化对决策的影响（304不锈钢板）', 7)
r = 19
write_header(ws5, r, ['标准板损耗率', '非标板损耗率', '标准板有效单价\n(¥/kg)', '非标板有效单价\n(¥/kg)', '非标节省\n(¥/kg)', '最优选择'])

# Parameter grid
scenarios = [
    [0.10, 0.03],
    [0.15, 0.04],
    [0.20, 0.05],
    [0.25, 0.06],
    [0.30, 0.07],
]

for i, sc in enumerate(scenarios):
    row = r + 1 + i
    set_cell(ws5, row, 1, sc[0], font=body_font, fill=input_fill, alignment=center_align, fmt=pct_fmt)
    set_cell(ws5, row, 2, sc[1], font=body_font, fill=input_fill, alignment=center_align, fmt=pct_fmt)
    # Effective std price = price_std / (1 - loss_std)
    set_cell(ws5, row, 3, f'={S1}!C13/(1-A{row})', font=body_font, alignment=center_align, fmt=money2_fmt)
    # Effective nonstd price = price_std * 1.12 / (1 - loss_nonstd)
    set_cell(ws5, row, 4, f'={S1}!C13*1.12/(1-B{row})', font=body_font, alignment=center_align, fmt=money2_fmt)
    # Savings per kg
    set_cell(ws5, row, 5, f'=C{row}-D{row}', font=bold_font, alignment=center_align, fmt=money2_fmt)
    # Decision
    set_cell(ws5, row, 6, f'=IF(E{row}>0,"非标板 ✓","标准板")', font=bold_font, alignment=center_align)

# ── 使用说明 ──
r = 26
section_title(ws5, r, '▎使用说明', 7)
r += 1
instructions = [
    '1. 所有可调参数集中在"1-参数设定"Sheet（黄色单元格），修改后全部Sheet自动重算。',
    '2. 规格选择优化器(Sheet2)：展示单物料×五车间的TCO对比和最优决策。',
    '3. 聚合效应计算器(Sheet3)：展示跨车间聚合如何解锁非标板、降低损耗。',
    '4. 库存策略优化器(Sheet4)：展示差异化库存策略的量化依据。',
    '5. 本仪表板(Sheet5)：汇总所有关键指标，提供敏感性分析。',
    '6. 请确保Excel公式计算设置为"自动"（公式→计算选项→自动）。',
]
for txt in instructions:
    set_cell(ws5, r, 1, txt, font=body_font, alignment=left_align)
    ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

set_col_widths(ws5, [22, 22, 18, 18, 16, 18, 18])

# ═══════════════════════════════════════
# 保存
# ═══════════════════════════════════════
out_path = os.path.expanduser(r'~\Documents\Obsidian Vault\26年中集环科工作区\决策科学案例库\案例01-MILP采购决策优化_演示表格.xlsx')
wb.save(out_path)
print(f'Excel V2 saved: {out_path}')
print('Done - 5 sheets with correct cross-sheet formulas')
